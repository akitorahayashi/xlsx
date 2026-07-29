# /// script
# requires-python = ">=3.12"
# dependencies = ["python-calamine==0.8.2", "openpyxl==3.1.5"]
# ///
"""
Read-only CLI over one .xlsx or .xlsm workbook. Every subcommand prints exactly
one JSON document on stdout.

The backend is chosen by what the caller asks for, because the cost difference
between backends is one to two orders of magnitude:

- overview: stdlib zipfile (parts, formula presence) + python-calamine (sheet
  dimensions, merged ranges, visibility). openpyxl is never imported here.
- rows / find: python-calamine by default; openpyxl only under --formulas.

Exit codes:
- 0: a result was produced.
- 1: the request was valid and the result is empty (find with no match, rows with
     no data rows).
- 2: a usage or runtime error. stderr carries one JSON document with "error" and
     an actionable "action", and stdout stays empty.
"""

from __future__ import annotations

import argparse
import csv
import datetime
import json
import posixpath
import re
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Iterator, NoReturn, Optional

import python_calamine as calamine

MAIN_NS = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
REL_NS = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}"
PKG_REL_NS = "{http://schemas.openxmlformats.org/package/2006/relationships}"

SUPPORTED_SUFFIXES = (".xlsx", ".xlsm")


class CLIError(Exception):
    """An input or runtime error carrying an action for the caller to fix."""

    def __init__(self, message: str, action: str) -> None:
        super().__init__(message)
        self.action = action


# --------------------------------------------------------------------------- #
# Shared helpers
# --------------------------------------------------------------------------- #


def resolve_workbook(raw: str) -> Path:
    path = Path(raw)
    if not path.exists():
        raise CLIError(f"File not found: {raw}", "Pass a path to an existing .xlsx or .xlsm workbook.")
    if path.suffix.lower() not in SUPPORTED_SUFFIXES:
        raise CLIError(
            f"Unsupported extension: {path.suffix or '(none)'}",
            "This skill reads .xlsx and .xlsm only. Convert other formats to .xlsx first.",
        )
    return path


def column_letter(index0: int) -> str:
    """Convert a zero-based column index to its A1 letters (0 -> 'A')."""
    n = index0 + 1
    letters = ""
    while n:
        n, remainder = divmod(n - 1, 26)
        letters = chr(65 + remainder) + letters
    return letters


def column_index(letters: str) -> int:
    """Convert A1 column letters to a one-based column number ('A' -> 1)."""
    n = 0
    for char in letters.upper():
        n = n * 26 + (ord(char) - 64)
    return n


@dataclass(frozen=True)
class CellRange:
    min_row: int
    min_col: int
    max_row: int
    max_col: int
    text: str


def parse_range(raw: str) -> CellRange:
    match = re.fullmatch(r"([A-Za-z]+)(\d+):([A-Za-z]+)(\d+)", raw.strip())
    if not match:
        raise CLIError(f"Invalid range: {raw!r}", "Use A1 notation with a colon, e.g. A1:F200.")
    c1, r1, c2, r2 = match.groups()
    col1, col2 = column_index(c1), column_index(c2)
    row1, row2 = int(r1), int(r2)
    return CellRange(
        min_row=min(row1, row2),
        min_col=min(col1, col2),
        max_row=max(row1, row2),
        max_col=max(col1, col2),
        text=raw.strip(),
    )


def is_empty(value: Any) -> bool:
    return value is None or value == ""


def to_json_value(value: Any) -> Any:
    """Map a backend cell value to its JSON representation.

    calamine reports every number as a float and an empty cell as ''. openpyxl
    reports an empty cell as None. Dates, datetimes, and times become ISO-8601
    strings; durations become seconds. Numbers and booleans keep their JSON types.
    """
    if is_empty(value):
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, datetime.datetime):
        return value.isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, datetime.time):
        return value.isoformat()
    if isinstance(value, datetime.timedelta):
        return value.total_seconds()
    if isinstance(value, (int, float, str)):
        return value
    return str(value)


def match_text(value: Any) -> Optional[str]:
    """The string a value is matched against, or None when the cell is empty."""
    rendered = to_json_value(value)
    if rendered is None:
        return None
    if isinstance(rendered, bool):
        return "TRUE" if rendered else "FALSE"
    return str(rendered)


def emit(document: dict[str, Any]) -> None:
    print(json.dumps(document, indent=2, ensure_ascii=False))


# --------------------------------------------------------------------------- #
# overview
# --------------------------------------------------------------------------- #


def count_parts(path: Path) -> dict[str, int]:
    with zipfile.ZipFile(path) as archive:
        names = archive.namelist()

    def count(pattern: Callable[[str], bool]) -> int:
        return sum(1 for name in names if pattern(name))

    return {
        "pivotTables": count(lambda n: bool(re.fullmatch(r"xl/pivotTables/pivotTable\d+\.xml", n))),
        "charts": count(lambda n: bool(re.fullmatch(r"xl/charts/chart\d+\.xml", n))),
        "images": count(lambda n: n.startswith("xl/media/") and not n.endswith("/")),
        # legacy comments appear as xl/comments1.xml (Excel) or xl/comments/comment1.xml (openpyxl)
        "comments": count(
            lambda n: bool(re.fullmatch(r"xl/comments\d*\.xml", n) or re.fullmatch(r"xl/comments/comment\d+\.xml", n))
        ),
        "threadedComments": count(lambda n: bool(re.fullmatch(r"xl/threadedComments/threadedComment\d+\.xml", n))),
    }


# A cell formula is the element <f> or <f ...>. The trailing byte disambiguates it
# from <formula1>/<formula2> (data validation, conditional formatting), which also
# begin with "<f" but are not cell formulas.
_FORMULA_TERMINATORS = frozenset(b" \t\r\n>/")


def _scan_for_formula(archive: zipfile.ZipFile, part: str, names: set[str]) -> bool:
    if part not in names:
        return False
    with archive.open(part) as stream:
        tail = b""
        while True:
            chunk = stream.read(1 << 20)
            if not chunk:
                return False
            buffer = tail + chunk
            start = 0
            while True:
                index = buffer.find(b"<f", start)
                if index == -1:
                    break
                after = index + 2
                if after >= len(buffer):
                    break  # marker at the buffer edge; re-check once the next chunk arrives
                if buffer[after] in _FORMULA_TERMINATORS:
                    return True
                start = after
            tail = buffer[-2:]  # carry the marker length across the boundary


def sheet_formula_presence(path: Path) -> dict[str, bool]:
    """Map each sheet name to whether its part contains a formula element.

    The sheet name to part path is resolved through xl/workbook.xml and
    xl/_rels/workbook.xml.rels, then the part is scanned in chunks for the first
    cell formula element.
    """
    import xml.etree.ElementTree as ET

    presence: dict[str, bool] = {}
    with zipfile.ZipFile(path) as archive:
        names = set(archive.namelist())
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))

        target_by_id: dict[str, str] = {}
        for rel in relationships.iter(f"{PKG_REL_NS}Relationship"):
            rel_id = rel.get("Id")
            target = rel.get("Target")
            if rel_id is not None and target is not None:
                target_by_id[rel_id] = target

        for sheet in workbook.iter(f"{MAIN_NS}sheet"):
            name = sheet.get("name")
            rel_id = sheet.get(f"{REL_NS}id")
            if name is None:
                continue
            target = target_by_id.get(rel_id) if rel_id is not None else None
            if target is None:
                presence[name] = False
                continue
            if target.startswith("/"):
                part = target.lstrip("/")
            else:
                part = posixpath.normpath(posixpath.join("xl", target))
            presence[name] = _scan_for_formula(archive, part, names)
    return presence


def cmd_overview(args: argparse.Namespace) -> int:
    path = resolve_workbook(args.file)
    parts = count_parts(path)
    formulas = sheet_formula_presence(path)

    workbook = calamine.CalamineWorkbook.from_path(str(path))
    sheets: list[dict[str, Any]] = []
    for meta in workbook.sheets_metadata:
        name = meta.name
        visible = meta.visible == calamine.SheetVisibleEnum.Visible
        try:
            sheet = workbook.get_sheet_by_name(name)
            start = sheet.start
        except Exception:
            start = None
            sheet = None

        if sheet is None or start is None:
            sheets.append(
                {
                    "name": name,
                    "visible": visible,
                    "empty": True,
                    "rows": 0,
                    "columns": 0,
                    "usedRange": None,
                    "mergedRanges": [],
                    "hasFormulas": formulas.get(name, False),
                }
            )
            continue

        start_row, start_col = start  # zero-based first used cell
        end_row = sheet.total_height  # zero-based last used row (== sheet.end[0])
        end_col = sheet.total_width  # zero-based last used column (== sheet.end[1])
        merged = [
            f"{column_letter(c0)}{r0 + 1}:{column_letter(c1)}{r1 + 1}"
            for (r0, c0), (r1, c1) in (sheet.merged_cell_ranges or [])
        ]
        sheets.append(
            {
                "name": name,
                "visible": visible,
                "empty": False,
                "rows": end_row - start_row + 1,
                "columns": end_col - start_col + 1,
                "usedRange": f"{column_letter(start_col)}{start_row + 1}:{column_letter(end_col)}{end_row + 1}",
                "mergedRanges": merged,
                "hasFormulas": formulas.get(name, False),
            }
        )

    emit({"file": str(path), "parts": parts, "sheets": sheets})
    return 0


# --------------------------------------------------------------------------- #
# rows
# --------------------------------------------------------------------------- #


def require_sheet(name: str, available: list[str]) -> None:
    if name not in available:
        raise CLIError(
            f"Sheet not found: {name!r}",
            "Pass one of the available sheet names: " + ", ".join(repr(s) for s in available) + ".",
        )


def require_header_row(header_row: int, last_row: Optional[int]) -> None:
    """Bound-check --header-row identically on both backends.

    ``last_row`` is the one-based last used row, or None when the backend cannot
    report it; the bounds are named in the message only when they are known.
    """
    if header_row >= 1 and (last_row is None or header_row <= last_row):
        return
    bounds = f" (1..{last_row})" if last_row is not None else ""
    raise CLIError(
        f"Header row {header_row} is outside the sheet{bounds}.",
        "Pass a --header-row within the sheet's used range.",
    )


def build_fields(header_cells: list[Any]) -> list[str]:
    fields: list[str] = []
    seen: set[str] = set()
    for value in header_cells:
        rendered = to_json_value(value)
        if rendered is None or (isinstance(rendered, str) and rendered.strip() == ""):
            raise CLIError(
                "Header row contains an empty field name.",
                "Choose a --header-row whose cells are all filled, or omit --header-row.",
            )
        name = str(rendered)
        if name in seen:
            raise CLIError(
                f"Header row repeats the field name {name!r}.",
                "Choose a --header-row with unique field names, or omit --header-row.",
            )
        seen.add(name)
        fields.append(name)
    return fields


def _calamine_row_iter(sheet: calamine.CalamineSheet, c0: int, c1: int) -> Iterator[tuple[int, list[Any]]]:
    """Yield (one-based row, values sliced to columns [c0, c1]) lazily.

    calamine's iter_rows is anchored at row 1 but offset to the first used column,
    so each row is re-anchored to column A before slicing.
    """
    start = sheet.start
    if start is None:
        return
    start_col = start[1]
    for index, row in enumerate(sheet.iter_rows()):
        end_col = start_col + len(row)
        yield index + 1, [row[c - start_col] if start_col <= c < end_col else "" for c in range(c0, c1 + 1)]


def _calamine_header_cells(sheet: calamine.CalamineSheet, header_row: int, c0: int, c1: int) -> list[Any]:
    grid = sheet.to_python(skip_empty_area=False, nrows=header_row)
    row = grid[header_row - 1] if 0 <= header_row - 1 < len(grid) else []
    return [row[c] if c < len(row) else "" for c in range(c0, c1 + 1)]


def _openpyxl_row_iter(worksheet: Any, min_col: int, max_col: Optional[int]) -> Iterator[tuple[int, list[Any]]]:
    for index, row in enumerate(worksheet.iter_rows(min_row=1, min_col=min_col, max_col=max_col, values_only=True)):
        yield index + 1, list(row)


def _openpyxl_header_cells(worksheet: Any, header_row: int, min_col: int, max_col: Optional[int]) -> list[Any]:
    rows = worksheet.iter_rows(
        min_row=header_row, max_row=header_row, min_col=min_col, max_col=max_col, values_only=True
    )
    return list(next(iter(rows), ()))


def _consume_rows(
    row_iter: Iterator[tuple[int, list[Any]]],
    r_lo: int,
    r_hi: Optional[int],
    header_row: Optional[int],
    keep_empty: bool,
    max_rows: Optional[int],
) -> tuple[list[list[Any]], int, bool]:
    """Apply row bounds, header exclusion, empty-row omission, and --max-rows.

    The iterator is consumed lazily and abandoned after one qualifying row beyond
    --max-rows, so a bounded request never materializes the whole sheet.
    """
    emitted: list[list[Any]] = []
    omitted = 0
    truncated = False
    for abs_row, values in row_iter:
        if header_row is not None and abs_row == header_row:
            continue
        if abs_row < r_lo:
            continue
        if r_hi is not None and abs_row > r_hi:
            break
        if all(is_empty(value) for value in values):
            if not keep_empty:
                omitted += 1
                continue
        if max_rows is not None and len(emitted) >= max_rows:
            truncated = True
            break
        emitted.append(values)
    return emitted, omitted, truncated


def _calamine_extract(
    path: Path,
    sheet_name: str,
    rng: Optional[CellRange],
    header_row: Optional[int],
    keep_empty: bool,
    max_rows: Optional[int],
) -> tuple[Optional[list[Any]], list[list[Any]], int, bool]:
    """Return (header cells, emitted rows, omitted count, truncated) for calamine."""
    workbook = calamine.CalamineWorkbook.from_path(str(path))
    require_sheet(sheet_name, workbook.sheet_names)
    sheet = workbook.get_sheet_by_name(sheet_name)

    if sheet.start is None:
        return None, [], 0, False

    last_row = sheet.total_height  # zero-based
    last_col = sheet.total_width  # zero-based

    if rng is not None:
        c0, c1 = rng.min_col - 1, rng.max_col - 1
        r_lo, r_hi = rng.min_row, rng.max_row
    else:
        c0, c1 = 0, last_col
        r_lo, r_hi = 1, None

    header_cells: Optional[list[Any]] = None
    if header_row is not None:
        require_header_row(header_row, last_row + 1)
        header_cells = _calamine_header_cells(sheet, header_row, c0, c1)

    emitted, omitted, truncated = _consume_rows(
        _calamine_row_iter(sheet, c0, c1), r_lo, r_hi, header_row, keep_empty, max_rows
    )
    return header_cells, emitted, omitted, truncated


def _openpyxl_extract(
    path: Path,
    sheet_name: str,
    rng: Optional[CellRange],
    header_row: Optional[int],
    keep_empty: bool,
    max_rows: Optional[int],
) -> tuple[Optional[list[Any]], list[list[Any]], int, bool]:
    """Return (header cells, emitted rows, omitted count, truncated) for openpyxl.

    read_only=True streams rows so a bounded request stays cheap; data_only=False
    keeps formula sources.
    """
    from openpyxl import load_workbook

    workbook = load_workbook(str(path), read_only=True, data_only=False)
    try:
        require_sheet(sheet_name, list(workbook.sheetnames))
        worksheet = workbook[sheet_name]

        sheet_max_row = worksheet.max_row
        if rng is not None:
            min_col, max_col = rng.min_col, rng.max_col
            r_lo, r_hi = rng.min_row, rng.max_row
        else:
            min_col, max_col = 1, worksheet.max_column
            r_lo, r_hi = 1, None

        header_cells: Optional[list[Any]] = None
        if header_row is not None:
            require_header_row(header_row, sheet_max_row)
            header_cells = _openpyxl_header_cells(worksheet, header_row, min_col, max_col)

        emitted, omitted, truncated = _consume_rows(
            _openpyxl_row_iter(worksheet, min_col, max_col), r_lo, r_hi, header_row, keep_empty, max_rows
        )
        return header_cells, emitted, omitted, truncated
    finally:
        workbook.close()


def cmd_rows(args: argparse.Namespace) -> int:
    path = resolve_workbook(args.file)
    rng = parse_range(args.range) if args.range else None
    header_row: Optional[int] = args.header_row
    backend = "openpyxl" if args.formulas else "calamine"

    extract = _openpyxl_extract if args.formulas else _calamine_extract
    header_cells, emitted, omitted, truncated = extract(
        path, args.sheet, rng, header_row, args.keep_empty_rows, args.max_rows
    )

    fields = build_fields(header_cells) if header_cells is not None else None

    if fields is not None:
        rows_out: list[Any] = [
            {name: to_json_value(value) for name, value in zip(fields, values)} for values in emitted
        ]
    else:
        rows_out = [[to_json_value(value) for value in values] for values in emitted]

    exit_code = 0 if emitted else 1

    if args.format in ("csv", "tsv"):
        delimiter = "," if args.format == "csv" else "\t"
        writer = csv.writer(sys.stdout, delimiter=delimiter, lineterminator="\n")
        if fields is not None:
            writer.writerow(fields)
        for values in emitted:
            writer.writerow(["" if to_json_value(v) is None else to_json_value(v) for v in values])
        if omitted:
            print(f"note: {omitted} empty row(s) omitted", file=sys.stderr)
        if truncated:
            print(f"note: output truncated at --max-rows {args.max_rows}", file=sys.stderr)
        return exit_code

    emit(
        {
            "file": str(path),
            "sheet": args.sheet,
            "backend": backend,
            "range": rng.text if rng else None,
            "headerRow": header_row,
            "fields": fields,
            "rowCount": len(emitted),
            "emptyRowsOmitted": omitted,
            "truncated": truncated,
            "rows": rows_out,
        }
    )
    return exit_code


# --------------------------------------------------------------------------- #
# find
# --------------------------------------------------------------------------- #


def build_matcher(args: argparse.Namespace) -> Callable[[str], bool]:
    if args.regex:
        try:
            pattern = re.compile(args.pattern, 0 if args.case_sensitive else re.IGNORECASE)
        except re.error as error:
            raise CLIError(
                f"Invalid regex: {error}", "Fix the regular expression, or drop --regex for substring search."
            )
        return lambda text: pattern.search(text) is not None
    if args.case_sensitive:
        needle = args.pattern
        return lambda text: needle in text
    needle = args.pattern.lower()
    return lambda text: needle in text.lower()


def _find_calamine(
    workbook: calamine.CalamineWorkbook,
    sheet_names: list[str],
    test: Callable[[str], bool],
    max_matches: Optional[int],
) -> tuple[list[dict[str, Any]], bool]:
    matches: list[dict[str, Any]] = []
    for name in sheet_names:
        sheet = workbook.get_sheet_by_name(name)
        if sheet.start is None:  # guard: calamine panics on iter of an empty sheet
            continue
        grid = sheet.to_python(skip_empty_area=False)
        for r, row in enumerate(grid):
            for c, value in enumerate(row):
                text = match_text(value)
                if text is None or not test(text):
                    continue
                matches.append({"sheet": name, "cell": f"{column_letter(c)}{r + 1}", "value": to_json_value(value)})
                if max_matches is not None and len(matches) >= max_matches:
                    return matches, True
    return matches, False


def _find_openpyxl(
    workbook: Any, sheet_names: list[str], test: Callable[[str], bool], max_matches: Optional[int]
) -> tuple[list[dict[str, Any]], bool]:
    matches: list[dict[str, Any]] = []
    for name in sheet_names:
        worksheet = workbook[name]
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                text = match_text(cell.value)
                if text is None or not test(text):
                    continue
                matches.append({"sheet": name, "cell": cell.coordinate, "value": to_json_value(cell.value)})
                if max_matches is not None and len(matches) >= max_matches:
                    return matches, True
    return matches, False


def select_sheets(requested: Optional[str], available: list[str]) -> list[str]:
    if requested is None:
        return available
    require_sheet(requested, available)
    return [requested]


def cmd_find(args: argparse.Namespace) -> int:
    """Search one workbook, opening the chosen backend's workbook exactly once.

    Sheet-name validation reads the names off the same open workbook the search
    then walks. A separate probe load cost a second full openpyxl parse of the
    file under --formulas, roughly doubling the wall time of every such call.
    """
    path = resolve_workbook(args.file)
    test = build_matcher(args)
    backend = "openpyxl" if args.formulas else "calamine"

    if args.formulas:
        from openpyxl import load_workbook

        workbook = load_workbook(str(path), read_only=True, data_only=False)
        try:
            sheet_names = select_sheets(args.sheet, list(workbook.sheetnames))
            matches, truncated = _find_openpyxl(workbook, sheet_names, test, args.max_matches)
        finally:
            workbook.close()
    else:
        calamine_workbook = calamine.CalamineWorkbook.from_path(str(path))
        sheet_names = select_sheets(args.sheet, calamine_workbook.sheet_names)
        matches, truncated = _find_calamine(calamine_workbook, sheet_names, test, args.max_matches)

    emit(
        {
            "file": str(path),
            "pattern": args.pattern,
            "backend": backend,
            "matchCount": len(matches),
            "truncated": truncated,
            "matches": matches,
        }
    )
    return 0 if matches else 1


# --------------------------------------------------------------------------- #
# argument parsing
# --------------------------------------------------------------------------- #


class JSONArgumentParser(argparse.ArgumentParser):
    """ArgumentParser that reports usage errors as the documented JSON on stderr.

    Subparsers inherit this class, so a missing positional, a malformed integer, or
    an invalid choice exits 2 with {"error", "action"} rather than plain usage text.
    """

    def error(self, message: str) -> NoReturn:
        print(
            json.dumps(
                {
                    "error": f"Argument error: {message}",
                    "action": "Fix the command's arguments; run the subcommand with --help for usage.",
                },
                ensure_ascii=False,
            ),
            file=sys.stderr,
        )
        raise SystemExit(2)


def positive_int(raw: str) -> int:
    """An argparse type for the cap flags, which have no meaning below 1.

    Without it the two caps disagreed at the boundary: --max-rows 0 emitted no
    rows with truncated set, while --max-matches 0 emitted one match.
    """
    value = int(raw)
    if value < 1:
        raise argparse.ArgumentTypeError("must be a positive integer")
    return value


def build_parser() -> argparse.ArgumentParser:
    parser = JSONArgumentParser(description="Inspect, extract from, and search one .xlsx or .xlsm workbook.")
    subparsers = parser.add_subparsers(dest="command", required=True)

    overview = subparsers.add_parser("overview", help="Structure, part inventory, and formula presence per sheet.")
    overview.add_argument("file")
    overview.set_defaults(func=cmd_overview)

    rows = subparsers.add_parser("rows", help="Extract a bounded window of rows from one sheet.")
    rows.add_argument("file")
    rows.add_argument("sheet")
    rows.add_argument("--range", help="A1 range bound, e.g. A1:F200.")
    rows.add_argument("--header-row", type=int, help="One-based row supplying field names.")
    rows.add_argument("--max-rows", type=positive_int, help="Cap emitted rows and set truncated.")
    rows.add_argument("--keep-empty-rows", action="store_true", help="Retain fully empty rows.")
    rows.add_argument("--formulas", action="store_true", help="Use openpyxl and emit formula sources.")
    rows.add_argument("--format", choices=("json", "csv", "tsv"), default="json")
    rows.set_defaults(func=cmd_rows)

    find = subparsers.add_parser("find", help="Search cell values, or formula sources under --formulas.")
    find.add_argument("file")
    find.add_argument("pattern")
    find.add_argument("--sheet", help="Restrict the search to one sheet.")
    find.add_argument("--regex", action="store_true", help="Treat the pattern as a regular expression.")
    find.add_argument("--case-sensitive", action="store_true", help="Match case exactly.")
    find.add_argument("--formulas", action="store_true", help="Use openpyxl and search formula sources.")
    find.add_argument("--max-matches", type=positive_int, help="Cap matches and set truncated.")
    find.set_defaults(func=cmd_find)

    return parser


def main(argv: Optional[list[str]] = None) -> int:
    args = build_parser().parse_args(argv)
    try:
        return int(args.func(args))
    except CLIError as error:
        print(json.dumps({"error": str(error), "action": error.action}, ensure_ascii=False), file=sys.stderr)
        return 2
    except Exception as error:  # noqa: BLE001 - surface any runtime failure as exit 2, not a traceback on stdout
        print(
            json.dumps(
                {
                    "error": f"{type(error).__name__}: {error}",
                    "action": "Check the file path, arguments, and that the workbook is a valid .xlsx/.xlsm.",
                },
                ensure_ascii=False,
            ),
            file=sys.stderr,
        )
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
