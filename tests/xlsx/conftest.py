"""Fixture workbooks built with openpyxl from the development virtualenv.

No test depends on the sample workbooks under .mx/; every fixture is constructed
into tmp_path.
"""

from __future__ import annotations

import datetime as dt
import zipfile
from pathlib import Path

import openpyxl
import pytest
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation

VBA_PART = "xl/vbaProject.bin"
VBA_STUB = b"stub-vba-project-part" * 4
VBA_CONTENT_TYPE = f'<Override PartName="/{VBA_PART}" ContentType="application/vnd.ms-office.vbaProject"/>'


@pytest.fixture
def workbook(tmp_path: Path) -> Path:
    """A workbook covering the read contract.

    Ledger has a header row, typed values including a date and a datetime,
    integer formulas, a merged range, frozen panes, and trailing empty rows, plus
    a chart. Empty is an entirely empty sheet (the calamine panic regression).
    Hidden is a hidden sheet with one value.
    """
    wb = openpyxl.Workbook()
    ledger = wb.active
    ledger.title = "Ledger"
    ledger["A1"], ledger["B1"], ledger["C1"] = "Date", "Item", "Qty"
    ledger["A2"], ledger["B2"], ledger["C2"] = dt.date(2026, 1, 5), "Widget", 3
    ledger["A3"], ledger["B3"], ledger["C3"] = dt.datetime(2026, 1, 6, 13, 30), "Gadget", 4
    ledger["C4"] = "=SUM(C2:C3)"
    # rows 5 and 6 are left empty on purpose
    ledger.merge_cells("A7:B7")
    ledger["A7"] = "Total"
    ledger.freeze_panes = "A2"

    chart = BarChart()
    chart.add_data(Reference(ledger, min_col=3, min_row=1, max_row=3), titles_from_data=True)
    ledger.add_chart(chart, "E1")

    wb.create_sheet("Empty")
    hidden = wb.create_sheet("Hidden")
    hidden.sheet_state = "hidden"
    hidden["A1"] = "secret"

    path = tmp_path / "book.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def header_workbook(tmp_path: Path) -> Path:
    """A workbook whose header rows exercise the two invalid-header cases."""
    wb = openpyxl.Workbook()
    empty_cell = wb.active
    empty_cell.title = "EmptyHdr"
    empty_cell["A1"], empty_cell["B1"], empty_cell["C1"] = "Date", None, "Qty"
    empty_cell["A2"], empty_cell["B2"], empty_cell["C2"] = 1, 2, 3

    duplicate = wb.create_sheet("DupHdr")
    duplicate["A1"], duplicate["B1"] = "Qty", "Qty"
    duplicate["A2"], duplicate["B2"] = 1, 2

    path = tmp_path / "headers.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def feature_workbook(tmp_path: Path) -> Path:
    """Sheets that exercise part counting, formula-presence, and offset dimensions.

    Commented carries a real cell comment. Validated has only a list data
    validation (a `<formula1>` element that must not be read as a cell formula).
    Offset's single populated cell is C5, so its used range does not start at A1.
    """
    wb = openpyxl.Workbook()
    commented = wb.active
    commented.title = "Commented"
    commented["A1"] = "hi"
    commented["A1"].comment = Comment("a note", "author")

    validated = wb.create_sheet("Validated")
    validation = DataValidation(type="list", formula1='"a,b,c"')
    validated.add_data_validation(validation)
    validation.add("B2")
    validated["A1"] = "plain"

    offset = wb.create_sheet("Offset")
    offset["C5"] = "only"

    path = tmp_path / "features.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def macro_workbook(tmp_path: Path) -> Path:
    """A .xlsm carrying a stub xl/vbaProject.bin part.

    openpyxl cannot author a VBA project, but keep_vba copies the part opaquely
    without parsing it, so stub bytes are enough to observe whether the round trip
    preserves it. The workbook is repacked from an openpyxl-built .xlsx in tmp_path
    with the part and its content-type Override added, so no binary is committed.
    """
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Macro"
    sheet["A1"] = 1
    plain = tmp_path / "plain.xlsx"
    wb.save(plain)

    path = tmp_path / "macro.xlsm"
    with zipfile.ZipFile(plain) as source, zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "[Content_Types].xml":
                data = data.decode().replace("</Types>", f"{VBA_CONTENT_TYPE}</Types>").encode()
            target.writestr(item, data)
        target.writestr(VBA_PART, VBA_STUB)
    return path
